import openpyxl
from openpyxl.styles import PatternFill

# Laad het Excel-bestand
pad_naar_excel = '"C:\Users\MukiLukunaITSynergy\IT Synergy\Aurora Group B.V. - Documents\General\Professional Services\Project - Online Werkplek\Inventarisatie Fase\KLANT_Onboarding_Checklist_V2024.xlsx"'  # Vervang dit door het pad naar jouw Excel-bestand
werkboek = openpyxl.load_workbook(pad_naar_excel)
werkblad = werkboek.active  # Of specificeer de naam van het werkblad met werkboek['Werkbladnaam']

# Lijst met waarden die rood gemarkeerd moeten worden
te_markeren_waarden = [
    "VERLT5CD103JKCG", "VERLT5CD1075XM6", "VERLT5CD1075XVW", "VERLT5CD1075XWZ",
    "VERLT5CD115NPH0", "VERLT5CD115NPKJ", "VERLT5CD115NPL0", "VERLT5CD1422GZS",
    "VERLT5CD1422H4Q", "VERLT5CD150BHYQ", "VERLT5CD203GK2S", "VERLT5CD203GK3P",
    "VERLT5CD203GK3Y", "VERLT5CD203GK55", "VERLT5CD206HJ85", "VERLT5CD2219QF3",
    "VERLT5CD2434NQC", "VERLT5CD2434NWV", "VERLT5CD2434NZ3", "VHM-LP44",
    "VHM5CD2064TSB"
]

# Definieer de rode opvulkleur
rode_opvulling = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Itereer door de cellen in kolom G (7e kolom)
for rij in werkblad.iter_rows(min_row=1, max_row=werkblad.max_row, min_col=7, max_col=7):
    for cel in rij:
        if cel.value in te_markeren_waarden:
            cel.fill = rode_opvulling

# Sla het gewijzigde Excel-bestand op
werkboek.save('pad/naar/jouw_bestand_met_opmaak.xlsx')  # Vervang dit door het gewenste pad en bestandsnaam
